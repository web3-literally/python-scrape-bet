from PyQt5.QtWidgets import QApplication
from PyQt5.QtWidgets import QApplication, QTableWidget, QTableWidgetItem, QHeaderView, QCheckBox, QMessageBox, QWidget, \
    QPushButton, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QFileDialog, QMainWindow
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtCore import QTimer

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException

import os
import sys
import threading
import time
import threading
import subprocess

import datetime

#import xlwt
#from xlwt import Workbook
import openpyxl
from openpyxl.worksheet import worksheet
from openpyxl.workbook import workbook
from openpyxl import cell
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment



############################
threads = {}
rank_results = []
IsFindingRelatedASIN = 0  # 0 => NONE, 1 => PROCESSING, 2 => FINISHED
related_ASINS = []


############################
class Scraping:
    def __init__(self, ASIN, Keyword, thread_name):
        self.ASIN = ASIN
        self.Keyword = Keyword
        self.thread_name = thread_name
        self.thread_flag = False
        self.thread_instance = None

        self.driver = None

    def start_thread(self):
        print("Start thread: " + self.thread_name + " ...")
        self.thread_flag = True
        threads[self.thread_name] = threading.Thread(target=self.scraping)
        threads[self.thread_name].start()

    def stop_thread(self):
        print("Stop thread: " + self.thread_name + " ...")
        self.thread_flag = False
        try:
            del threads[self.thread_name]
            print('--Stopped--')
        except:
            print('--Stop Failed--')
            pass

    def get_product_url(self):
        return "https://www.amazon.com/gp/product/" + self.ASIN

    def get_keyword_rul(self):
        strKeywordUrl = ""
        words = self.Keyword.split(' ')
        for word in words:
            if (strKeywordUrl == ""):
                strKeywordUrl += word
            else:
                strKeywordUrl += "+" + word
        return "https://www.amazon.com/s?k=" + strKeywordUrl

    def get_related_ASINS(self):
        print('get_related_ASINS')
        global IsFindingRelatedASIN

        if IsFindingRelatedASIN == 0:
            related_ASINS.append(self.ASIN)
            IsFindingRelatedASIN = 1
            self.driver.get(self.get_product_url())
            time.sleep(3)
            color_Items = self.driver.find_elements_by_xpath("//li[@data-defaultasin]")
            for Item in color_Items:
                related_ASINS.append(Item.get_attribute('data-defaultasin'))
            IsFindingRelatedASIN = 2

    def get_rank_from_keyword(self):
        print('get_rank_from_keyword : ' + self.get_keyword_rul())
        while IsFindingRelatedASIN == 1:
            time.sleep(3)

        self.driver.get(self.get_keyword_rul())
        time.sleep(3)
        page_num = 1
        while 1:
            try:
                # product_div = ''
                # product_index = "10000"
                # for ASIN_NODE in related_ASINS:
                #     try:
                #         product_div_temp = self.driver.find_element_by_xpath("//div[@data-asin='" + ASIN_NODE + "']")
                #         product_index_temp = product_div_temp.get_attribute("data-index")
                #         if int(product_index) > int(product_index_temp):
                #             product_div = product_div_temp
                #             product_index = product_index_temp
                #     except:
                #         continue
                # if product_div == '':
                #     raise
                # rank = {}
                # rank['keyword'] = self.Keyword
                # rank['page'] = str(page_num)
                # rank['position'] = product_index
                # rank_results.append(rank)

                all_products = self.driver.find_elements_by_xpath("//div[@data-index]")
                product_index = 1
                Sponsored_Index = 1
                for product in all_products:
                    product_asin = product.get_attribute('data-asin')
                    if product_asin == '':
                        continue
                    else:
                        if related_ASINS.__contains__(product_asin):
                            rank = {'keyword': self.Keyword,
                                    'organic_page': str(page_num),
                                    'organic_position': str(product_index),
                                    'sponsored_page': str(page_num),
                                    'sponsored_position': str(Sponsored_Index)}
                            rank_results.append(rank)
                            print("--Detected Ranking!--")
                            break

                        try:
                            elem = product.find_element_by_css_selector('.a-row.a-spacing-micro')
                            if elem.find_element_by_css_selector('.a-size-base.a-color-secondary').text == 'Sponsored':
                                Sponsored_Index += 1
                            else:
                                product_index += 1
                        except:
                            product_index += 1

                print("ranking end!")
                break
            except:
                try:
                    next_button = self.driver.find_element_by_class_name("a-last")
                except:
                    break
                is_disabled = "a-disabled" in next_button.get_attribute("class")
                if is_disabled:
                    break
                next_button.click()
                time.sleep(3)
                page_num += 1

    def scraping(self):
        chrome_path = os.path.join(os.path.dirname(__file__), 'drivers', 'chromedriver')
        chrome_options1 = Options()
        chrome_options1.add_argument("--incognito")
        chrome_options1.add_argument("--headless")
        self.driver = webdriver.Chrome(executable_path=chrome_path, options=chrome_options1)
        time.sleep(3)

        self.get_related_ASINS()
        self.get_rank_from_keyword()

        self.driver.quit()
        time.sleep(3)
        self.stop_thread()


class MainWnd(QMainWindow):
    def __init__(self, parent=None):
        super(MainWnd, self).__init__(parent)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.on_Timer)

        # Implememt GUI
        layout = QVBoxLayout()

        lay_hor = QHBoxLayout()
        self.edit_FilePath = QLineEdit()
        self.btn_FilePath = QPushButton('  Keyword File Open  ')
        self.btn_FilePath.clicked.connect(self.on_btn_FilePath_Clicked)

        lay_hor.addWidget(self.edit_FilePath)
        lay_hor.addWidget(self.btn_FilePath)

        lay_hor_0 = QHBoxLayout()
        self.lbl_ASIN = QLabel()
        self.lbl_Keywords = QLabel()

        lay_hor_0.addWidget(QLabel('ASIN: '))
        lay_hor_0.addWidget(self.lbl_ASIN)
        lay_hor_0.addWidget(QLabel('Keywords'))
        lay_hor_0.addWidget(self.lbl_Keywords)

        lay_hor_1 = QHBoxLayout()
        self.btn_Scan = QPushButton("SCAN")
        self.btn_Save = QPushButton("SAVE")
        self.btn_Close = QPushButton("CLOSE")

        self.btn_Scan.clicked.connect(self.on_btn_Scan_Clicked)
        self.btn_Save.clicked.connect(self.on_btn_Save_Clicked)
        self.btn_Close.clicked.connect(self.on_btn_Close_Clicked)

        lay_hor_1.addWidget(self.btn_Scan)
        lay_hor_1.addWidget(self.btn_Save)
        lay_hor_1.addWidget(self.btn_Close)

        self.table_detail = QTableWidget()
        self.table_detail.setColumnCount(5)
        self.table_detail.setRowCount(3)
        header = self.table_detail.horizontalHeader()
        for i in range(self.table_detail.columnCount()):
            header.setSectionResizeMode(i, QHeaderView.Stretch)
        self.table_detail.horizontalHeader().setVisible(False)
        self.table_detail.verticalHeader().setVisible(False)

        self.table_detail.setSpan(0, 0, 1, 5)
        newItem = QTableWidgetItem('ASIN: ')
        self.table_detail.setItem(0, 0, newItem)
        self.table_detail.setSpan(1, 0, 2, 1)
        newItem = QTableWidgetItem('Keyword')
        self.table_detail.setItem(1, 0, newItem)
        self.table_detail.setSpan(1, 1, 1, 2)
        newItem = QTableWidgetItem('Organic Position')
        self.table_detail.setItem(1, 1, newItem)
        self.table_detail.setSpan(1, 3, 1, 2)
        newItem = QTableWidgetItem('Sponsored Position')
        self.table_detail.setItem(1, 3, newItem)
        newItem = QTableWidgetItem('Page')
        self.table_detail.setItem(2, 1, newItem)
        newItem = QTableWidgetItem('Position')
        self.table_detail.setItem(2, 2, newItem)
        newItem = QTableWidgetItem('Page')
        self.table_detail.setItem(2, 3, newItem)
        newItem = QTableWidgetItem('Position')
        self.table_detail.setItem(2, 4, newItem)

        layout.addLayout(lay_hor)
        layout.addLayout(lay_hor_0)
        layout.addLayout(lay_hor_1)
        layout.addWidget(self.table_detail)

        self.statusBar().showMessage('Ready')
        centralWnd = QWidget()
        centralWnd.setLayout(layout)
        self.setCentralWidget(centralWnd)
        self.setFixedSize(1000, 800)
        self.setWindowTitle("AMAZON PRODUCTS ANALYZE")

        # Variables
        self.ASIN = ''
        self.Keywords = []

    def on_btn_FilePath_Clicked(self):
        print('on_btn_FilePath_Clicked :')
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self, "Open Keyword File", "",
                                                  "All Files (*);;Keyword Files (*.txt)", options=options)
        if fileName:
            print(fileName)
            self.edit_FilePath.setText(fileName)
            f = open(fileName, 'r')
            if f.mode == 'r':
                contents = f.read()
                self.ASIN = fileName.split('/')[-1].split('.')[0]
                self.Keywords = contents.split(',')
                print('ASIN: ' + self.ASIN)
                print('Keywords: ' + contents)
                self.lbl_ASIN.setText(self.ASIN)
                self.lbl_Keywords.setText(contents)
                self.statusBar().showMessage('Keyword File Open Success!')

    def on_btn_Scan_Clicked(self):
        rank_results.clear()
        threads.clear()
        IsFindingRelatedASIN = 0
        related_ASINS.clear()

        thread_index = 1
        for word in self.Keywords:
            thread_name = "thread_" + str(thread_index)
            threads[thread_name] = Scraping(self.ASIN, word, thread_name)
            threads[thread_name].start_thread()
            thread_index += 1

        if thread_index == 1:
            return

        time.sleep(3)
        self.timer.start(5)

        self.statusBar().showMessage('Scanning.....')
        self.btn_FilePath.setEnabled(False)
        self.btn_Scan.setEnabled(False)
        self.btn_Save.setEnabled(False)
        self.btn_Close.setEnabled(False)
    def on_btn_Save_Clicked(self):
      
        dest_filename = r'xx.xlsx'
        wb=openpyxl.load_workbook(dest_filename)
        ws = wb.worksheets[0]
        # add column headings. NB. these must be strings
        ############ Unmerging Header ###########
        print(ws['B1'].value)
        col=2
        print(openpyxl.utils.cell.get_column_letter(col))
        while ws.cell(1,col).value!='' and ws.cell(1,col).value!=None:
            if ws.cell(1,col).value!='' and ws.cell(1,col).value!=None:        
                ws.unmerge_cells( openpyxl.utils.cell.get_column_letter(col)+'1:'+openpyxl.utils.cell.get_column_letter(col+3)+'1')
                ws.unmerge_cells( openpyxl.utils.cell.get_column_letter(col)+'2:'+openpyxl.utils.cell.get_column_letter(col+1)+'2')
                ws.unmerge_cells( openpyxl.utils.cell.get_column_letter(col+2)+'2:'+openpyxl.utils.cell.get_column_letter(col+3)+'2')        
                col = col+4
        row=4     
        col=1
        for rank in rank_results:
                ws.cell(column=col, row=row, value=rank['keyword'])
                ws['A'+ str(row)].alignment = Alignment(horizontal="center", vertical="center")
                row= row+1
        ws.insert_cols(2)
        col=2        
        row=4     
        for rank in rank_results:
                ws.cell(column=col, row=row, value=rank['organic_page'])
                ws['B'+ str(row)].alignment = Alignment(horizontal="center", vertical="center")
                row= row+1
        ws.insert_cols(3)
        col=3
        row=4
        for row in range(4, 20):     
                ws.cell(column=col, row=row, value=9)
                ws['C'+ str(row)].alignment = Alignment(horizontal="center", vertical="center")
                row=row+1
        ws.insert_cols(4)
        col=4
        row=4
        for row in range(4, 20):     
                ws.cell(column=col, row=row, value=8)
                ws['D'+ str(row)].alignment = Alignment(horizontal="center", vertical="center")
                row=row+1
        ws.insert_cols(5)
        col=5
        row=4
        for row in range(3, 20):     
                ws.cell(column=col, row=row, value=7)         
                ws['E'+ str(row)].alignment = Alignment(horizontal="center", vertical="center")
                row=row+1
        ws.cell(column=2,row=3,value='Page')
        ws.cell(column=3,row=3,value='Position')
        ws.cell(column=4,row=3,value='Page')
        ws.cell(column=5,row=3,value='Position')
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.merge_cells('B1:E1')
        ws.merge_cells('B2:C2')
        ws.merge_cells('D2:E2')

        curTime = datetime.datetime.now().strftime('%m/%d/%Y')

        ws.cell(column=2,row=1,value='Today:'+curTime)

        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(column=2,row=2,value='Organic Position')
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(column=4,row=2,value='Sponsord Position')
        ws["D2"].alignment = Alignment(horizontal="center", vertical="center")

        if ws['F1'].value!='' and ws['F1'].value!=None:
            #ws.merge_cells('F1:I1')
            #ws.merge_cells('F2:G2')
            #ws.merge_cells('H2:I2')
            ws.cell(column=6,row=1,value=ws['F1'].value[6:])
        col=6
        while ws.cell(1,col).value!='' and ws.cell(1,col).value!=None:
            if ws.cell(1,col).value!='' and ws.cell(1,col).value!=None:
                ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = 12
                ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col+1)].width = 12
                ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col+2)].width = 12
                ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col+3)].width = 12        
                ws.merge_cells( openpyxl.utils.cell.get_column_letter(col)+'1:'+openpyxl.utils.cell.get_column_letter(col+3)+'1')
                ws.merge_cells( openpyxl.utils.cell.get_column_letter(col)+'2:'+openpyxl.utils.cell.get_column_letter(col+1)+'2')
                ws.merge_cells( openpyxl.utils.cell.get_column_letter(col+2)+'2:'+openpyxl.utils.cell.get_column_letter(col+3)+'2')        
                col = col+4
        wb.save(filename=dest_filename)

    ############# Close #################
    def on_btn_Close_Clicked(self):
        sys.exit(1)

    def set_Table_Data(self, ASIN, Keyword, Organic_Page, Organic_Position, Sponsored_Page, Sponsored_Position):
        numRows = self.table_detail.rowCount()
        self.table_detail.setItem(0, 0, QTableWidgetItem("ASIN: " + ASIN))
        self.table_detail.insertRow(numRows)
        self.table_detail.setItem(numRows, 0, QTableWidgetItem(Keyword))
        self.table_detail.setItem(numRows, 1, QTableWidgetItem(Organic_Page))
        self.table_detail.setItem(numRows, 2, QTableWidgetItem(Organic_Position))
        self.table_detail.setItem(numRows, 3, QTableWidgetItem(Sponsored_Page))
        self.table_detail.setItem(numRows, 4, QTableWidgetItem(Sponsored_Position))

    def on_Timer(self):
        threadNames = threads.keys()
        if threadNames.__len__() == 0:
            self.timer.stop()
            try:
                for rank in rank_results:
                    self.set_Table_Data(self.ASIN, rank['keyword'], rank['organic_page'], rank['organic_position'],
                                        rank['sponsored_page'], rank['sponsored_position'])
            except:
                print('No Result')

            self.statusBar().showMessage('Scanning Finished!')
            self.btn_FilePath.setEnabled(True)
            self.btn_Scan.setEnabled(True)
            self.btn_Save.setEnabled(True)
            self.btn_Close.setEnabled(True)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWnd()
    window.show()
    sys.exit(app.exec())
